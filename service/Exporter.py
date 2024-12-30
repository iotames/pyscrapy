import PIL
from openpyxl import Workbook #, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from . import Config
import os
import time, mimetypes
from openpyxl.drawing.image import Image
from utils.crypto import get_md5
from openpyxl.utils import get_column_letter

# 手动添加 .webp 文件扩展名的 MIME 类型映射
mimetypes.add_type('image/webp', '.webp')



# Exporter 数据导出器。
# xlsx支持导出图片。图片的路径应为相对路径。如 sitename/md5sum.jpg
class Exporter:

    wb: Workbook
    sheet: Worksheet
    output_file: str
    images_dir: str
    image_title = 'Thumbnail'
    rowheight = 100
    imagewidth = 100
    
    # 构造函数
    # filename参数为文件名。不包含文件后缀
    def __init__(self, filename='output', sheet_title='Sheet1'):
        cnf = Config.get_instance()
        self.images_dir = cnf.get_images_path()
        self.output_file = os.path.join(cnf.get_export_dir(), "{}_{}.xlsx".format(filename, time.strftime("%Y%m%d%H%M%S", time.localtime())))
        self.wb = Workbook()
        self.sheet = self.wb.create_sheet(index=0, title=sheet_title)        
        # self.sheet = self.wb.active
        # self.sheet.title = "Scraped Data"
        if os.path.isfile(self.output_file):
            raise RuntimeError("文件已存在，请先删除")
            # self.wb = load_workbook(self.output_file)
            # self.sheet = self.wb.worksheets[0]
    
    def set_image_title(self, title):
        self.image_title = title

    def append_row(self, row: list):
        self.sheet.append(list(row))
    
    def add_image(self, img: Image, column_index, row_index):
        img.anchor = get_column_letter(column_index) + str(row_index)
        if img is not None:
            self.sheet.add_image(img)

    def get_image_filepath_by_url(self, imgurl: str, dirname: str) -> str:
        file_ext = ".jpg"
        filename = get_md5(imgurl) + file_ext
        # file_ext = mimetypes.guess_extension(mimetypes.guess_type(url)[0])
        return self.get_imagepath_by_filename(filename, dirname)

    def get_imagepath_by_filename(self, filename: str, dirname: str) -> Image:
        return os.path.join(self.images_dir, dirname, filename)
    
    def get_image_by_url(self, imgurl: str, dirname: str) -> Image:
        return self.get_image_by_filepath(self.get_image_filepath_by_url(imgurl, dirname))

    def get_image_by_filepath(self, fpath: str) -> Image:
        try:
            if not os.path.isfile(fpath):
                raise ValueError("图片{}不存在".format(fpath))
            image = Image(fpath)
            image.height = self.rowheight
            image.width = self.imagewidth
            # image.anchor = get_column_letter(start_col) + str(row_index)
            # print(image.anchor)
            return image
        except PIL.UnidentifiedImageError as e:
            raise e

    def save(self):
        self.sheet.column_dimensions['A'].width = 13
        for row in self.sheet.iter_rows(min_row=1, max_row=self.sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                self.sheet.row_dimensions[cell.row].height = self.rowheight
        self.wb.save(self.output_file)

    def to_xlsx(self):
        sheet = self.sheet
        # sheet.sheet_format.defaultRowHeight = 30
        title_row = ('商品ID', 'SPU', '图片', '分类', '商品标题', '商品链接', '更新时间', '价格/$', '状态')
        title_col = 1
        for title in title_row:
            sheet.cell(1, title_col, title)
            title_col += 1
        goods_list = []
        goods_row_index = 2
        for goods in goods_list:
            goods_col_index = 1
            time_str = self.timestamp_to_str(goods.updated_at, "%Y-%m-%d %H:%M")
            # 商品信息元组
            image = self.get_image_info(goods.local_image) if goods.local_image else ''
            goods_info_list = [
                goods.id, goods.asin, image, goods.category_name, goods.title, time_str, goods.price
            ]
            print(goods_info_list)
            # 返回商品信息递增列 next col index
            self.set_values_to_row(sheet, goods_info_list, goods_row_index, goods_col_index)
            goods_row_index += 1
        self.wb.save(self.output_file)

    def get_image_info(self, path: str) -> dict:
        if not path.startswith('/'):
            path = self.images_dir + os.path.sep + path
        if not os.path.isfile(path):
            return {'type': str, 'path': path}
        image = {
            'type': Image,
            'path': path,
            'size': (100, 100)
        }
        return image

    @staticmethod
    def set_values_to_row(sheet: Worksheet, values_list: list, row_index, start_col=1):
        for cell_value in values_list:
            if isinstance(cell_value, dict):
                if cell_value['type'] == Image:
                    # print('===========================================Image===' + cell_value['path'])
                    try:
                        image = Image(cell_value['path'])
                        image.width, image.height = cell_value['size']
                        image.anchor = get_column_letter(start_col) + str(row_index)
                        # print(image.anchor)
                        sheet.add_image(image)
                    except PIL.UnidentifiedImageError as e:
                        sheet.cell(row_index, start_col, '')

                if cell_value['type'] == str:
                    sheet.cell(row_index, start_col, cell_value['path'])
            else:
                sheet.cell(row_index, start_col, cell_value)
            start_col += 1
        return start_col

    @staticmethod
    def timestamp_to_str(timestamp=None, format_str="%Y-%m-%d %H:%M") -> str:
        time_tuple = time.localtime(timestamp)
        return time.strftime(format_str, time_tuple)


    @staticmethod
    def debug():
        exp = Exporter("debugexport1111")
        exp.append_row(['Thumnail', '2', '3', '4', '5', '6', '7', '8', '9'])
        exp.append_row(["", '2222', '3333', '4', '5', '6', '7', '8', '9'])
        exp.append_row(["", '2222', '3333', '4', '5', '6', '7', '8', '9'])
        imgs = [
            exp.get_imagepath_by_filename("0a5890653dd80b014a9a010deecd7ba2.jpg", "4tharq"),
            exp.get_imagepath_by_filename("0b023b2ae23b5af4dad335f7aba9e8f8.jpg", "4tharq")
        ]
        id = 2
        for img in imgs:
            exp.add_image(img, 1, id)
            id += 1
        exp.save()


if __name__ == '__main__':
    pass