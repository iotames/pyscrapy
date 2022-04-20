import PIL
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from Config import Config
from service.DB import DB
from pyscrapy.models import Site, Translator
import os
import shutil
from config.api_server import ApiServer
import time
from sqlalchemy.orm.session import Session
# from scrapy.utils.project import get_project_settings
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from pyscrapy.spiders.basespider import BaseSpider
# from translate import Translator
from service.Translator import Translator as TranslatorService


class BaseOutput:

    site_name: str
    site_id: int
    db_session: Session
    wb: Workbook
    work_sheet: Worksheet
    output_dir = Config.ROOT_PATH + '/runtime'
    output_file: str = output_dir + '/{}_' + time.strftime("%Y-%m-%d_%H_%M", time.localtime()) + '.xlsx'
    images_dir: str
    downloads_dirname = "downloads"
    download_filename = None
    # server_config: ApiServer
    trans_dic_en_zh = {}
    translator: TranslatorService

    @staticmethod
    def get_db_session() -> Session:
        db = DB.get_instance(Config().get_database())
        db.ROOT_PATH = Config.ROOT_PATH
        return db.get_db_session()

    def __init__(self, sheet_title='库存详情', filename='output'):

        self.translator = TranslatorService()  # Translator(to_lang='chinese', provider='mymemory')  # , proxies={'http': '127.0.0.1:1080'}

        # self.server_config = ApiServer()
        self.images_dir = BaseSpider.custom_settings.get('IMAGES_STORE')  # get_project_settings().get('IMAGES_STORE')
        
        self.db_session = self.get_db_session()
        # TODO 因文件名故，xlsx文件通常仅走新增路线
        self.output_file = self.output_file.format(filename)
        if os.path.isfile(self.output_file):
            self.wb = load_workbook(self.output_file)
            self.work_sheet = self.wb.worksheets[0]
        else:
            self.wb = Workbook()
            self.work_sheet = self.wb.create_sheet(index=0, title=sheet_title)
        site = self.db_session.query(Site).filter(Site.name == self.site_name).first()
        if not site:
            raise RuntimeError(self.site_name + " 在数据库中不存在")
        self.site_id = site.id

    def to_chinese(self, content: str, check_local_dict=True) -> str:
        # return self.translator.translate(content)
        # must use self.get_trans_dic_all() before
        if len(content) < 5:
            return content
        if check_local_dict and content in self.trans_dic_en_zh:
            return self.trans_dic_en_zh[content]
        cn_content = self.translator.to_chinese(content)
        if check_local_dict:
            self.add_trans_dict(content, cn_content)
        return cn_content

    def get_trans_dic_all(self):
        dict_list = self.db_session.query(Translator).filter(Translator.trans_type == 0).all()
        if len(dict_list) > 0:
            for dt in dict_list:
                self.trans_dic_en_zh[dt.from_lang] = dt.to_lang

    def add_trans_dict(self, fro: str, to: str, trans_type=0):
        if trans_type > 0:
            raise ValueError("trans_type value must == 0")
        self.trans_dic_en_zh[fro] = to
        self.db_session.add(Translator(from_lang=fro, to_lang=to, trans_type=trans_type))
        self.db_session.commit()

    def get_trans_value(self, fro: str, trans_type=0):
        if trans_type > 0:
            raise ValueError("trans_type value must == 0")
        if fro in self.trans_dic_en_zh:
            return self.trans_dic_en_zh[fro]
        return None

    def get_image_info(self, path: str) -> dict:
        if not path.startswith('/'):
            path = self.images_dir + os.path.sep + path
        if not os.path.isfile(path):
            return {'type': str, 'path': path}
        image = {
            'type': Image,
            'path': path,
            'size': (100, 100)
        }
        return image

    @staticmethod
    def set_values_to_row(sheet: Worksheet, values_list: list, row_index, start_col=1):
        for cell_value in values_list:
            if isinstance(cell_value, dict):
                if cell_value['type'] == Image:
                    # print('===========================================Image===' + cell_value['path'])
                    try:
                        image = Image(cell_value['path'])
                        image.width, image.height = cell_value['size']
                        image.anchor = get_column_letter(start_col) + str(row_index)
                        # print(image.anchor)
                        sheet.add_image(image)
                    except PIL.UnidentifiedImageError as e:
                        sheet.cell(row_index, start_col, '')

                if cell_value['type'] == str:
                    sheet.cell(row_index, start_col, cell_value['path'])
            else:
                sheet.cell(row_index, start_col, cell_value)
            start_col += 1
        return start_col

    @staticmethod
    def timestamp_to_str(timestamp=None, format_str="%Y-%m-%d %H:%M") -> str:
        time_tuple = time.localtime(timestamp)
        return time.strftime(format_str, time_tuple)

    def copy_to_download_path(self, srcfile):
        if not os.path.isfile(srcfile):
            msg = "file: {} not exists!".format(srcfile)
            raise RuntimeError(msg)
        shutil.copyfile(srcfile, self.get_downloads_dir() + os.path.sep + self.download_filename)

    def is_download_file_exists(self) -> bool:
        filepath = self.get_downloads_dir() + os.path.sep + self.download_filename
        if os.path.isfile(filepath):
            print("file exist: " + filepath)
            return True
        return False

    # def get_downloads_dir(self):
    #     downloads_dir = self.server_config.get_config()['static_folder'] + os.path.sep + self.downloads_dirname
    #     if not os.path.exists(downloads_dir):
    #         os.makedirs(downloads_dir)
    #     return downloads_dir

    def output(self):
        pass

    # def output_to_excel(self):
        # sheet = self.work_sheet
        # # sheet.sheet_format.defaultRowHeight = 30
        # title_row = ('商品ID', '分类名', '商品名', '商品链接', '商品状态', '更新时间', '规格1', '规格2', 'SKU名', '价格', '库存')
        # title_col = 1
        # for title in title_row:
        #     sheet.cell(1, title_col, title)
        #     title_col += 1
        # goods_list = self.db_session.query(Goods).filter(Goods.site_id == self.site_id).all()
        # sku_row_index = 2
        #
        # for goods in goods_list:
        #     goods_col_index = 1
        #     start_row_index = sku_row_index
        #     time_str = self.timestamp_to_str(goods.updated_at)
        #     # 商品信息元组
        #     goods_info_list = (goods.id, goods.category_name, goods.title, goods.url, Goods.statuses_map[goods.status], time_str)
        #     # 返回商品信息递增列 next col index
        #     goods_col_index = self.set_values_to_row(sheet, goods_info_list, sku_row_index, goods_col_index)

        # self.wb.save(self.output_file)


if __name__ == '__main__':
    output = BaseOutput()
    # output.output_to_excel()
