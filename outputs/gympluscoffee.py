import json

from pyscrapy.models import Goods, GoodsSku, GoodsCategory
from outputs.baseoutput import BaseOutput
import time
from translate import Translator
from openpyxl import load_workbook


class GympluscoffeeOutput(BaseOutput):

    site_name = 'gympluscoffee'
    translator: Translator
    categories: list
    goods_model_list: list
    origin_file = ''
    inventory_cols_len = 1  # 新生成的文件库存栏数量

    def __init__(self):
        super(GympluscoffeeOutput, self).__init__('SKU库存详情', self.site_name)
        self.categories = self.db_session.query(GoodsCategory).all()
        self.translator = Translator(to_lang='chinese', provider='mymemory')

    def get_parent_by_category_id(self, category_id: int):
        parent_id = 0
        model = None
        for category in self.categories:
            if category.id == category_id:
                parent_id = category.parent_id
                break
        for category in self.categories:
            if category.id == parent_id:
                model = category
        return model

    def output_to_excel(self):
        sheet = self.work_sheet
        sheet.sheet_format.defaultRowHeight = 100
        # sheet.sheet_format.defaultColWidth = 100

        origin_wb = load_workbook(self.origin_file)
        origin_sheet = origin_wb.worksheets[0]
        title_ext = []
        row_index = 1
        inventory_map = {}
        for row_data in origin_sheet:
            if row_index == 1:
                for i in range(22, 21 + self.inventory_cols_len):
                    print(row_data[i].value)
                    title_ext.append(row_data[i].value)
                row_index += 1
                continue
            # row_data[0] 索引从0开始 | 20 sku名 | 22 库存
            if row_data[22].value == "---":
                row_index += 1
                continue
            url = row_data[5].value
            if not url:
                row_index += 1
                continue
            sku_text = row_data[20].value if row_data[20].value else ''
            inv_key = url + sku_text
            if inv_key not in inventory_map:
                inventory_values = []
                for i in range(22, 21 + self.inventory_cols_len):
                    inventory_values.append(row_data[i].value)
                inventory_map[inv_key] = inventory_values
            row_index += 1

        title_row = ['商品ID', '图片', '上级分类', '分类名', '商品标题', '商品链接', '商品状态', '更新时间',
                     '评论数', '价格/EUR', '商品简介', '商品详情', '织物布料',
                     '5星评论', '4星评论', '3星评论', '2星评论', '1星评论',
                     '规格1', '规格2', 'SKU名', '价格']
        title_row.extend(title_ext)
        title_row.append('库存{}'.format(time.strftime("%m_%d", time.localtime())))

        title_col = 1
        for title in title_row:
            sheet.cell(1, title_col, title)
            title_col += 1
        self.goods_model_list = self.db_session.query(Goods).filter(Goods.site_id == self.site_id).all()

        print('total goods model = ' + str(len(self.goods_model_list)))
        # TODO 新版本使用，不兼容就版本。需要导入旧数据
        goods_urls = []
        unique_goods_list = []
        for model in self.goods_model_list:
            if model.url not in goods_urls:
                goods_urls.append(model.url)
                unique_goods_list.append(model)
        print('total goods model = ' + str(len(unique_goods_list)))
        sku_row_index = 2

        # for goods in self.goods_model_list:
        for goods in unique_goods_list:
            goods_col_index = 1
            start_row_index = sku_row_index
            time_tuple = time.localtime(goods.updated_at)
            time_str = time.strftime("%Y-%m-%d %H:%M", time_tuple)
            reviews_num = goods.reviews_num
            details = goods.details
            price = goods.price
            schema = ''
            detail = ''
            fabric = ''
            rating5, rating4, rating3, rating2, rating1 = (0, 0, 0, 0, 0)
            if details:
                details = json.loads(details)
                schema = details['schema']  # + " 翻译： " + self.to_chinese(details['schema'])
                detail = details['details']  # + " 翻译： " + self.to_chinese(details['details'])
                fabric = details['fabric']  # + " 翻译： " + self.to_chinese(details['fabric'])
                if reviews_num > 0:
                    rating = details['rating']
                    rating5 = rating['5']
                    rating4 = rating['4']
                    rating3 = rating['3']
                    rating2 = rating['2']
                    rating1 = rating['1']
            # 商品信息元组
            image = ''
            if goods.local_image:
                image = self.get_image_info(goods.local_image)
            p_category_name = ''
            p_category = self.get_parent_by_category_id(goods.category_id)
            if p_category:
                p_category_name = p_category.name
            goods_info_list = [goods.id, p_category_name, goods.category_name, goods.title, goods.url,
                               Goods.statuses_map[goods.status], time_str,
                               reviews_num, price, schema, detail, fabric, rating5, rating4, rating3, rating2, rating1]
            goods_row_info = goods_info_list.copy()
            goods_row_info.insert(1, image)
            # print(goods_row_info)
            # 返回商品信息递增列 next col index
            goods_col_index = self.set_values_to_row(sheet, goods_row_info, sku_row_index, goods_col_index)

            goods_sku_list = self.db_session.query(GoodsSku).filter(
                GoodsSku.site_id == self.site_id, GoodsSku.goods_id == goods.id).all()
            sku_len = len(goods_sku_list)
            for sku in goods_sku_list:
                if sku_row_index > start_row_index:
                    # goods的SKU数量大于1， 从当前EXCEL行数量另起一行，复制goods基本信息填入。
                    goods_col_index = 1
                    if sku.local_image:
                        image = self.get_image_info(sku.local_image)
                    # 从第2个SKU开始，从goods拷贝基本信息
                    sku_row_info = goods_info_list.copy()
                    sku_row_info.insert(1, image)
                    print(sku_row_info)
                    # 从第2个SKU开始，都要拷贝goods基本信息写入EXCEL
                    goods_col_index = self.set_values_to_row(sheet, sku_row_info, sku_row_index, goods_col_index)
                price = format(sku.price/100, '.2f')

                sku_text = sku.title if sku.title else ''
                inventory_key = goods.url + sku_text
                if inventory_key in inventory_map:
                    inventory_values = inventory_map[inventory_key]
                else:
                    inventory_values = []
                    for i in range(1, self.inventory_cols_len):
                        inventory_values.append('')
                inventory_values.append(sku.inventory_quantity)
                sku_info_list = [sku.option1, sku.option2, sku_text, price]
                sku_info_list.extend(inventory_values)

                # SKU信息列紧接商品信息列之后
                sku_col_index = goods_col_index
                for sku_info in sku_info_list:
                    sheet.cell(sku_row_index, sku_col_index, sku_info)
                    # SKU信息列递增
                    sku_col_index += 1
                sku_row_index += 1

            if sku_len == 0:
                sku_row_index += 1

        self.wb.save(self.output_file)
        # new_wb = load_workbook(self.output_file)
        # new_sheet = new_wb.worksheets[0]


if __name__ == '__main__':
    gc = GympluscoffeeOutput()
    gc.inventory_cols_len = 14  # 新的文件库存栏数量
    gc.origin_file = gc.output_dir + '/gympluscoffee_2022-02-07_10_22.xlsx'
    gc.output_to_excel()
