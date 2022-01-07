from pyscrapy.models import Goods, RankingGoods, RankingLog, GoodsReview, SpiderRunLog, GroupLog, GroupGoods
from outputs.baseoutput import BaseOutput
from pyscrapy.extracts.amazon import Common as XAmazon, GoodsReviews as XReviews
import json
from sqlalchemy import and_
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from time import time
from pyscrapy.enum.spider import *


class AmazonOutput(BaseOutput):

    site_name = NAME_AMAZON
    child: str
    cell_fill = PatternFill("solid", fgColor="1874CD")
    goods_id_to_title = {}
    colors_show_times = {}
    group_name = ""
    ranking_log_model: RankingLog
    group_log_model: GroupLog
    
    def __init__(self, run_log: SpiderRunLog):
        self.child = run_log.spider_child
        link_id = run_log.link_id
        query_args = {"id": int(link_id)}
        if self.child == CHILD_GOODS_REVIEWS_BY_GROUP:
            self.group_log_model = GroupLog.get_model(GroupGoods.get_db_session(), query_args)
            self.group_name = self.group_log_model.code

        if self.child == CHILD_GOODS_REVIEWS_BY_RANKING:
            self.ranking_log_model = RankingLog.get_model(RankingLog.get_db_session(), query_args)
            self.group_name = self.ranking_log_model.category_name

        filename = "{}_{}".format(self.site_name, self.group_name).replace(' ', '_').replace("'", "-").replace("&", "and")
        self.download_filename = "{}_{}.xlsx".format(filename, link_id)
        super(AmazonOutput, self).__init__('商品信息列表', filename)

    @property
    def log_model(self):
        if self.child == CHILD_GOODS_REVIEWS_BY_GROUP:
            return self.group_log_model
        if self.child == CHILD_GOODS_REVIEWS_BY_RANKING:
            return self.ranking_log_model

    def set_colors_show_times(self, colorr: str):
        if colorr in self.colors_show_times:
            self.colors_show_times[colorr] = self.colors_show_times[colorr] + 1
        else:
            self.colors_show_times[colorr] = 1

    def set_reviews_colors(self, reviews: list, detail_rows, color_partial=True):
        for review in reviews:
            product_title = self.goods_id_to_title[str(review.goods_id)]
            review_title = review.title
            rating = review.rating_value
            sku_text = review.sku_text
            url = review.url
            color_text = review.color
            if sku_text and not review.color:
                color_text = XReviews.get_color_in_sku_text(sku_text)
            if not color_text:
                print(review.sku_text)
            row_detail = (review.time_str, product_title, review_title, review.body, color_text, rating, sku_text, url, 0)
            # print(row_detail)
            detail_rows.append(row_detail)
            if not color_text:
                continue
            print(color_text)
            if color_partial:
                colors = color_text.split(' ')
                for color in colors:
                    self.set_colors_show_times(color)
            else:
                color = color_text
                self.set_colors_show_times(color)

    def output(self):
        if self.is_download_file_exists():
            return True
        sheet = self.work_sheet

        if not self.log_model:
            raise RuntimeError('找不到排行榜数据')
        if time() - self.log_model.created_at > 3600 * 72:
            raise RuntimeError('最近排行榜数据已超过72小时, 请重新采集')

        db_session = self.db_session
        x_goods_list = []
        if self.child == CHILD_GOODS_REVIEWS_BY_RANKING:
            x_goods_list = db_session.query(RankingGoods).filter_by(**{'ranking_log_id': self.log_model.id}).order_by(
                RankingGoods.rank_num.asc()).all()
        if self.child == CHILD_GOODS_REVIEWS_BY_GROUP:
            x_goods_list = db_session.query(GroupGoods).filter_by(**{'group_log_id': self.log_model.id}).order_by(
                GroupGoods.rank_num.asc()).all()

        current_time = int(time())
        month_in_12 = current_time - 3600 * 24 * 365
        month_in_6 = current_time - 3600 * 24 * 180
        month_in_3 = current_time - 3600 * 24 * 90
        month_in_2 = current_time - 3600 * 24 * 60
        month_in_1 = current_time - 3600 * 24 * 30
        week_in_1 = current_time - 3600 * 24 * 7

        # sheet.sheet_format.defaultRowHeight = 30
        title_row = ('商品ID', 'code', '亚马逊ASIN(蓝色表示重复)', '分类', '图片', '商品标题', '商品链接', '上架时间', '数据爬取时间', '价格/US$', '原价', '节省',
                     '总评论数(包含不同颜色)', '1年内评论', '6个月内评论数', '3个月内评论数', '2个月内评论', '1个月内评论', '1周内评论', '大类排名', '当前排名', '商品描述', '所有排名')
        title_col = 1
        for title in title_row:
            sheet.cell(1, title_col, title)
            title_col += 1

        # goods_list = self.db_session.query(Goods).filter(
        #     Goods.site_id == self.site_id,
        #     # Goods.merchant_id == 1
        # ).all()
        print(len(x_goods_list))

        # ASIN 分组整理 , 并剔除重复的SPU。 START
        asin_list = []
        asin_goods_map = {}
        code_list = []
        for xgoods in x_goods_list:
            goods_model = Goods.get_model(db_session, {'id': xgoods.goods_id})
            asin = goods_model.asin
            code = XAmazon.get_code_by_goods_url(goods_model.url)
            goods_model.code = code  # 重写 code
            if asin not in asin_list:
                if code not in code_list:  # 剔除重复的SPU
                    asin_list.append(asin)
                    code_list.append(code)
                    asin_goods_map[asin] = [goods_model]
            else:
                if code not in code_list:  # 剔除重复的SPU
                    asin_goods_map[asin].append(goods_model)

        goods_list = []
        for asin, gd_list in asin_goods_map.items():
            for model in gd_list:
                goods_list.append(model)
        # ASIN 分组整理, 并剔除重复的SPU。 END

        goods_row_index = 2
        for goods_model in goods_list:
            self.goods_id_to_title[str(goods_model.id)] = goods_model.title
            goods_col_index = 1
            time_str = self.timestamp_to_str(goods_model.updated_at, "%Y-%m-%d %H:%M")
            # 商品信息元组
            image = ''
            if goods_model.local_image:
                image = self.get_image_info(goods_model.local_image)
            details = json.loads(goods_model.details)

            sale_at = details['sale_at'] if 'sale_at' in details else ''
            asin = details['asin']
            root_rank = details['root_rank']
            goods_url = goods_model.url

            has_group_list = [CHILD_GOODS_REVIEWS_BY_GROUP, CHILD_GOODS_REVIEWS_BY_RANKING]
            if self.child in has_group_list:
                category_name = self.group_name
            else:
                category_name = goods_model.category_name if goods_model.category_name else self.group_name

            rank_list = details['rank_list']
            rank_num = details['rank_num'] if 'rank_num' in details else 0
            rank_detail = ''
            for rank in rank_list:
                # AmazonSpider.get_site_url(rank['url'])
                rank_detail += rank['category_text'] + " : " + rank['rank_text'] + '\n|'
            details_items = ""
            for item in details['items']:
                details_items += item + "\n|"
            reviews_month_12 = db_session.query(GoodsReview).filter(
                GoodsReview.goods_id == goods_model.id, GoodsReview.review_time > month_in_12).count()
            reviews_month_6 = db_session.query(GoodsReview).filter(
                GoodsReview.goods_id == goods_model.id, GoodsReview.review_time > month_in_6).count()
            reviews_month_3 = db_session.query(GoodsReview).filter(
                GoodsReview.goods_id == goods_model.id, GoodsReview.review_time > month_in_3).count()
            reviews_month_2 = db_session.query(GoodsReview).filter(
                GoodsReview.goods_id == goods_model.id, GoodsReview.review_time > month_in_2).count()
            reviews_month_1 = db_session.query(GoodsReview).filter(
                GoodsReview.goods_id == goods_model.id, GoodsReview.review_time > month_in_1).count()
            reviews_week_1 = db_session.query(GoodsReview).filter(
                GoodsReview.goods_id == goods_model.id, GoodsReview.review_time > week_in_1).count()

            goods_info_list = [
                goods_model.id, goods_model.code, asin, category_name, image, goods_model.title, goods_url, sale_at, time_str,
                goods_model.price, details['price_base'], details['price_save'], goods_model.reviews_num, reviews_month_12, reviews_month_6,
                reviews_month_3, reviews_month_2, reviews_month_1, reviews_week_1, root_rank, rank_num, details_items,
                rank_detail
            ]
            print(goods_info_list)
            # 返回商品信息递增列 next col index
            self.set_values_to_row(sheet, goods_info_list, goods_row_index, goods_col_index)
            goods_row_index += 1

        # 标记重复的ASIN
        total_asin_list = []
        for row in sheet.rows:
            cell = row[2]
            asin = cell.value
            if asin in total_asin_list:
                cell.fill = self.cell_fill
                print(asin)
            else:
                total_asin_list.append(asin)

        sheet_reviews = self.wb.create_sheet(title='评论详情', index=1)
        details_title_row = ('时间', '商品', '评论概要', '评论详情', '颜色', '评分', 'SKU', '评论源地址', '尺码问题')
        detail_rows = [details_title_row]
        for goods_model in goods_list:
            reviews = self.db_session.query(GoodsReview).filter(
                and_(
                    GoodsReview.site_id == self.site_id,
                    GoodsReview.goods_id == goods_model.id,
                    GoodsReview.review_time > month_in_12
                )).order_by(GoodsReview.review_time.desc()).all()
            self.set_reviews_colors(reviews, detail_rows, False)

        for row in detail_rows:
            # 基础评论数据
            sheet_reviews.append(row)

        sheet_analysis = self.wb.create_sheet(title="颜色统计", index=2)
        sheet_analysis.append(('颜色', '出现次数'))
        for color_name, color_times in self.colors_show_times.items():
            # 热卖颜色分析
            sheet_analysis.append((color_name, color_times))
        size_index = len(details_title_row) - 1
        # sku_text_index = size_index - 2
        for row in sheet_reviews.rows:
            # 尺码问题标记
            row_index = 0
            for cell in row:
                if type(cell.value) == str and (row_index == 2 or row_index == 3):
                    if cell.value.find("size") > -1:
                        cell.fill = self.cell_fill
                        row[size_index].value = 1
                    if cell.value.find("Size") > -1:
                        cell.fill = self.cell_fill
                        row[size_index].value = 1
                row_index += 1

        self.wb.save(self.output_file)
        # TODO BUG： 通过API接口下载和通过本入口生成的EXCEL， 文件内容不一样。
        # self.copy_to_download_path(self.output_file)

    def update_excel(self, filepath):
        self.wb = load_workbook(filepath)
        self.work_sheet = self.wb.worksheets[0]
        row_index = 0
        for row in self.work_sheet.rows:
            print(row_index)
            if row_index == 0:
                row_index += 1
                continue
            url = row[6].value
            print(url)
            code = XAmazon.get_code_by_goods_url(url)
            print(code)
            row[1].value = code
            row_index += 1

        self.wb.save(filepath)

    def debug(self):
        print(self.group_name)


if __name__ == '__main__':
    db_session = SpiderRunLog.get_db_session()
    log = SpiderRunLog.get_model(db_session, {'id': 173})  # 164 165 168 169
    ot = AmazonOutput(log)
    ot.output()
