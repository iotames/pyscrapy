import time
from pyscrapy.models import Goods, RankingGoods, RankingLog, GoodsReview, SpiderRunLog
from outputs.baseoutput import BaseOutput
import json
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class SheinOutput(BaseOutput):
    site_name = 'shein'
    rank_category: str
    ranking_log_model: RankingLog

    current_time = int(time.time())
    month_in_12 = current_time - 3600 * 24 * 365
    month_in_6 = current_time - 3600 * 24 * 180
    month_in_3 = current_time - 3600 * 24 * 90
    month_in_2 = current_time - 3600 * 24 * 60
    month_in_1 = current_time - 3600 * 24 * 30
    week_in_1 = current_time - 3600 * 24 * 7

    def __init__(self, run_log: SpiderRunLog):
        ranking_log_id = int(run_log.link_id)
        self.ranking_log_model = RankingLog.get_model(RankingLog.get_db_session(), {"id": ranking_log_id})
        self.rank_category = self.ranking_log_model.category_name
        filename = "{}_{}".format(self.site_name, self.rank_category).replace(' ', '_').replace("'", "-").replace("&", "and")
        self.download_filename = "{}_{}.xlsx".format(filename, str(self.ranking_log_model.id))
        super(SheinOutput, self).__init__('商品信息列表', filename)
    
    @classmethod
    def get_goods_info_list_by_model(cls, goods: Goods, rank_num: int) -> list:
        details = json.loads(goods.details)
        first_at = details['first_review_time'] if 'first_review_time' in details else ''
        image = cls.get_image_info(goods.local_image) if goods.local_image else ''
        color = details['color'] if 'color' in details else ''
        updated_at = cls.timestamp_to_str(goods.updated_at)
        brand = details['brand'] if 'brand' in details else ''
        sku_num = 1
        if 'relation_colors' in details:
            relation_colors = details['relation_colors']
            if relation_colors:
                sku_num += len(relation_colors)
        db_session = cls.get_db_session()
        reviews_month_12 = db_session.query(GoodsReview).filter(
            GoodsReview.goods_spu == goods.asin, GoodsReview.review_time > cls.month_in_12).count()
        reviews_month_6 = db_session.query(GoodsReview).filter(
            GoodsReview.goods_spu == goods.asin, GoodsReview.review_time > cls.month_in_6).count()
        reviews_month_3 = db_session.query(GoodsReview).filter(
            GoodsReview.goods_spu == goods.asin, GoodsReview.review_time > cls.month_in_3).count()
        reviews_month_2 = db_session.query(GoodsReview).filter(
            GoodsReview.goods_spu == goods.asin, GoodsReview.review_time > cls.month_in_2).count()
        reviews_month_1 = db_session.query(GoodsReview).filter(
            GoodsReview.goods_spu == goods.asin, GoodsReview.review_time > cls.month_in_1).count()
        reviews_week_1 = db_session.query(GoodsReview).filter(
            GoodsReview.goods_spu == goods.asin, GoodsReview.review_time > cls.week_in_1).count()

        goods_info_list = [
            goods.id, image, rank_num, first_at, goods.code, goods.asin, color, goods.category_name,
            goods.title, goods.url, updated_at, goods.price, reviews_month_12, reviews_month_6, reviews_month_3,
            reviews_month_2, reviews_month_1, reviews_week_1, brand, sku_num, goods.reviews_num
        ]
        return goods_info_list

    @classmethod
    def output_reviews_by_goods_model(cls, goods: Goods):
        wb = Workbook()
        sheet = wb.create_sheet(index=0, title="goods_reviews")
        cls.set_sheet_title_rows(sheet)
        goods_info_list = cls.get_goods_info_list_by_model(goods, 0)
        cls.set_values_to_row(sheet, goods_info_list, 2, 1)
        db_session = cls.get_db_session()
        reviews = db_session.query(GoodsReview).filter(GoodsReview.goods_spu == goods.asin, GoodsReview.review_time > cls.month_in_12).all()
        for review in reviews:
            print(review)
        wb.save(cls.output_file.format("goods_reviews"))

    @staticmethod
    def set_sheet_title_rows(sheet: Worksheet):
        title_row = ('ID', '图片', '排名', '首评时间', 'goods_id', 'SPU', '颜色', '品类', '商品标题', '商品链接', '更新时间',
                     '价格/US$', '1年内评论', '6个月内评论数', '3个月内评论数', '2个月内评论', '1个月内评论', '1周内评论', '所属品牌', 'SKU数', '总评论数')
        title_col = 1
        for title in title_row:
            sheet.cell(1, title_col, title)
            title_col += 1

    def output_top_100(self):
        # if self.is_download_file_exists():
        #     return True
        sheet = self.work_sheet
        log = self.ranking_log_model
        if not log:
            raise RuntimeError('找不到排行榜数据')
        # if time.time() - log.created_at > 3600 * 240:
        #     raise RuntimeError('最近排行榜数据已超过240小时（10天）, 请重新采集')
        db_session = self.db_session
        ranking_goods_list = db_session.query(RankingGoods).filter_by(**{'ranking_log_id': log.id}).order_by(
            RankingGoods.rank_num.asc()).all()
        
        self.set_sheet_title_rows(sheet)

        goods_row_index = 2

        for rgoods in ranking_goods_list:
            goods = Goods.get_model(db_session, {'id': rgoods.goods_id})
            goods_info_list = self.get_goods_info_list_by_model(goods, rgoods.rank_num)
            # 返回商品信息递增列 next col index
            self.set_values_to_row(sheet, goods_info_list, goods_row_index, 1)
            goods_row_index += 1
        self.wb.save(self.output_file)
        # TODO BUG： 通过API接口下载和通过本入口生成的EXCEL， 文件内容不一样。
        # self.copy_to_download_path(self.output_file)

    def output(self):
        self.output_top_100()


if __name__ == '__main__':
    db_session = SpiderRunLog.get_db_session()
    log = SpiderRunLog.get_model(db_session, {'id': 35})  # 24 25
    ot = SheinOutput(log)
    ot.output()


